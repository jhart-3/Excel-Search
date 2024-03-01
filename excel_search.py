import sys

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import (QAbstractItemView, QAction, QApplication,
                             QCheckBox, QComboBox, QDialog, QFileDialog,
                             QGridLayout, QHBoxLayout, QLabel, QLineEdit,
                             QMainWindow, QMenu, QMessageBox, QPushButton,
                             QScrollArea, QStyledItemDelegate,
                             QStyleOptionViewItem, QTableWidget,
                             QTableWidgetItem, QVBoxLayout, QWidget)

class ExcelSettings:
    def __init__(self):
        self.sheet_name = ''
        self.search_column = ''
        self.additional_columns = []

class ExcelSettingsDialog(QDialog):
    settings_updated = pyqtSignal(ExcelSettings)

    def __init__(self, excel_settings, sheet_names, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Search Parameters')

        self.sheet_names = sheet_names

        self.sheet_name_input = QComboBox(self)
        self.sheet_name_input.addItems(list(self.sheet_names.keys()))
        self.sheet_name_input.setCurrentText(excel_settings.sheet_name)

        self.search_column_input = QComboBox(self)
        self.search_column_input.setCurrentText(excel_settings.search_column)

        self.additional_columns_input = QWidget(self)  # Container widget for additional columns
        self.additional_columns_scrollarea = QScrollArea()  # Scroll area for additional columns
        self.additional_columns_scrollarea.setWidgetResizable(True)
        self.additional_columns_scrollarea.setWidget(self.additional_columns_input)

        self.additional_columns_layout = QGridLayout(self.additional_columns_input)  # Use QGridLayout for checkboxes

        self.additional_columns_checkboxes = []  # List to store the checkboxes

        self.ok_button = QPushButton('OK', self)
        self.ok_button.clicked.connect(self.accept)
        self.ok_button.setFixedSize(160, 30)

        self.cancel_button = QPushButton('Cancel', self)
        self.cancel_button.clicked.connect(self.reject)
        self.cancel_button.setFixedSize(160, 30)

        self.button_layout = QHBoxLayout()
        self.button_layout.addWidget(self.ok_button, alignment=Qt.AlignCenter)
        self.button_layout.addWidget(self.cancel_button, alignment=Qt.AlignCenter)

        self.layout = QGridLayout(self)  # Use QGridLayout for the dialog layout
        self.layout.addWidget(QLabel('Sheet Name:'), 0, 0)
        self.layout.addWidget(self.sheet_name_input, 0, 1)
        self.layout.addWidget(QLabel('Search Column:'), 1, 0)
        self.layout.addWidget(self.search_column_input, 1, 1)
        self.layout.addWidget(QLabel('Additional Columns:'), 2, 0)
        self.layout.addWidget(self.additional_columns_scrollarea, 2, 1)
        self.layout.addLayout(self.button_layout, 3, 0, 1, 2, alignment=Qt.AlignCenter)

        self.sheet_name_input.currentTextChanged.connect(self.update_search_columns)
        self.search_column_input.currentTextChanged.connect(self.update_additional_columns)

        # Initially update the search columns and additional columns based on the selected sheet
        self.update_search_columns(self.sheet_name_input.currentText())
        self.update_additional_columns(self.search_column_input.currentText())

    def update_search_columns(self, sheet_name):
        self.search_column_input.clear()
        if sheet_name and sheet_name in self.sheet_names:
            columns = self.sheet_names[sheet_name].columns
            self.search_column_input.setEnabled(True)  # Enable the search column input if there are columns available
            self.search_column_input.addItems(columns)

        if not self.search_column_input.count():  # If no columns available, disable the search column input
            self.search_column_input.setEnabled(False)

        if not self.search_column_input.isEnabled() and self.sheet_name_input.count() == 1:
            # Only one sheet available, but no columns found
            self.search_column_input.setEnabled(True)
            self.search_column_input.addItem("No columns available")
            self.search_column_input.setCurrentIndex(0)

    def update_additional_columns(self, search_column):
        # Clear the layout and checkboxes
        for i in reversed(range(self.additional_columns_layout.count())):
            checkbox = self.additional_columns_layout.itemAt(i).widget()
            self.additional_columns_layout.removeWidget(checkbox)
            checkbox.deleteLater()

        self.additional_columns_checkboxes = []

        sheet_name = self.sheet_name_input.currentText()
        if sheet_name and sheet_name in self.sheet_names:
            columns = list(self.sheet_names[sheet_name].columns)
            if search_column in columns:
                columns.remove(search_column)

            if columns:
                row = 0
                col = 0

                for column in columns:
                    checkbox = QCheckBox(column)
                    self.additional_columns_checkboxes.append(checkbox)
                    self.additional_columns_layout.addWidget(checkbox, row, col)

                    col += 1
                    if col == 3:
                        col = 0
                        row += 1

        # Fix: Update the additional_columns_input widget to force the layout to update
        self.additional_columns_input.setLayout(self.additional_columns_layout)

    def accept(self):
        sheet_name = self.sheet_name_input.currentText()
        search_column = self.search_column_input.currentText()
        additional_columns = []
        for checkbox in self.additional_columns_checkboxes:
            if checkbox.isChecked():
                additional_columns.append(checkbox.text())

        if not sheet_name or not search_column:
            QMessageBox.warning(self, 'Warning', 'Sheet Name and Search Column are required.')
        else:
            new_settings = ExcelSettings()
            new_settings.sheet_name = sheet_name
            new_settings.search_column = search_column
            new_settings.additional_columns = additional_columns
            self.settings_updated.emit(new_settings)
            super().accept()

class CustomItemDelegate(QStyledItemDelegate):
    def initStyleOption(self, option, index):
        super().initStyleOption(option, index)
        option.displayAlignment = Qt.AlignCenter  # Set text alignment to center
        option.textElideMode = Qt.ElideNone  # Disable text eliding
        option.features |= QStyleOptionViewItem.WrapText  # Enable text wrapping

class SearchThread(QThread):
    search_complete = pyqtSignal(list)

    def __init__(self, excel_file_path, excel_settings, search_phrase):
        super().__init__()
        self.excel_file_path = excel_file_path
        self.excel_settings = excel_settings
        self.search_phrase = search_phrase

    def run(self):
        try:
            data = []

            workbook = load_workbook(filename=self.excel_file_path, read_only=True)
            if self.excel_settings.sheet_name in workbook.sheetnames:
                df = pd.read_excel(self.excel_file_path, sheet_name=self.excel_settings.sheet_name, keep_default_na=False)
                columns_to_search = [self.excel_settings.search_column] + self.excel_settings.additional_columns

                # Filter rows based on the search column and search phrase
                matched_rows = df[df[self.excel_settings.search_column].astype(str).str.contains(self.search_phrase, case=False, na=False, regex=True)]

                if not matched_rows.empty:
                    # Select the desired columns
                    selected_columns = [self.excel_settings.search_column] + columns_to_search
                    matched_rows = matched_rows[selected_columns]

                    # Add the sheet name to each row
                    matched_rows.insert(0, 'Sheet', self.excel_settings.sheet_name)

                    # Replace NaN and empty strings with None
                    matched_rows = matched_rows.replace({np.nan: None, '': None})

                    # Convert the DataFrame to a list of dictionaries
                    data = matched_rows.to_dict('records')

            self.search_complete.emit(data)
        except Exception as e:
            print(f"Search error: {str(e)}")
            self.search_complete.emit([])

class ExcelSearchApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Excel Search')
        self.resize(800, 600)
        self.setStyleSheet('''
            QMainWindow {
                background-color: #444444;
            }                       
            QWidget {
                background-color: #F0F0F0;
            }
            QLabel {
                color: #333333;
                font-size: 18px;
            }
            QLineEdit {
                background-color: #FFFFFF;
                border: 1px solid #AAAAAA;
                padding: 5px;
            }
            QPushButton {
                background-color: #007C78;
                color: #FFFFFF;
                padding: 5px 10px;
            }
            QTreeView {
                background-color: #FFFFFF;
                color: #333333;
                font-size: 16px;
                alternate-background-color: #F0F0F0;
                selection-background-color: #007C78;
            }
            QTreeView::item {
                padding: 5px;
            }
            QMenu {
                background-color: #FFFFFF;
                color: #333333;
                font-size: 16px;
                border: 1px solid #AAAAAA;
            }
            QMenu::item {
                padding: 5px;
            }
            QMenu::item:selected {
                background-color: #007C78;
                color: #FFFFFF;
            }
            QTextEdit {
                background-color: #FFFFFF;
                border: 1px solid #AAAAAA;
                padding: 5px;
                font-size: 16px;
            }
            QTableWidget {
                background-color: #FFFFFF;
                color: #333333;
                font-size: 16px;
            }
            QComboBox {
                background-color: #FFFFFF;
                border: 1px solid #AAAAAA;
                color: #333333;
                font-size: 14px;
                padding: 5px;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: center right;
                width: 20px;
                border-left-width: 1px;
                border-left-color: #AAAAAA;
                border-left-style: solid;
                background-color: #FFFFFF;
            }
            QComboBox::item {
                background-color: #FFFFFF;
                color: #333333;
            }
            QComboBox::item:selected {
                background-color: #007C78;
                color: #FFFFFF;
            }
            QComboBox::item:hover {
                background-color: #007C78;
                color: #FFFFFF;
            }
        ''')

        self.setup_ui()

        self.excel_file_path = ''
        self.excel_settings = ExcelSettings()
        self.search_thread = None
        self.search_results = []

    def setup_ui(self):
        QApplication.setStyle("Fusion")
        self.label_font = QFont('Sans-Serif', 12)
        self.button_font = QFont('Exo', 12)

        # Create the UI components
        self.search_label = QLabel('Enter search phrase:', self)
        self.search_label.setFont(self.label_font)
        self.search_label.setAlignment(Qt.AlignCenter)

        self.search_input = QLineEdit(self)
        self.search_input.returnPressed.connect(self.search_excel)
        self.search_input.setFont(self.label_font)
        self.search_input.setAlignment(Qt.AlignCenter)

        self.search_button = QPushButton('Search', self)
        self.search_button.setFont(self.button_font)
        self.search_button.clicked.connect(self.search_excel)

        self.choose_file_action = QAction('Choose Excel File', self)
        self.choose_file_action.triggered.connect(self.get_excel_file_path)

        self.settings_action = QAction('Search Parameters', self)
        self.settings_action.triggered.connect(self.open_settings)

        self.export_action = QAction('Export', self)
        self.export_action.triggered.connect(self.export_to_excel)

        self.menu = QMenu(self)
        self.menu.addAction(self.choose_file_action)
        self.menu.addAction(self.settings_action)
        self.menu.addAction(self.export_action)

        self.menu_button = QPushButton('Options', self)
        self.menu_button.setFont(self.button_font)
        self.menu_button.setMenu(self.menu)

        self.results_table = QTableWidget(self)
        self.results_table.setColumnCount(0)
        self.results_table.setRowCount(0)
        self.results_table.horizontalHeader().setStretchLastSection(True)
        self.results_table.verticalHeader().setVisible(False)
        self.results_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.results_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.results_table.setItemDelegate(CustomItemDelegate())
        self.results_table.setSortingEnabled(True)  # Enable sorting by clicking on the table headers
        self.results_table.setSelectionMode(QAbstractItemView.ExtendedSelection)

        self.search_layout = QHBoxLayout()
        self.search_layout.addWidget(self.search_label)
        self.search_layout.addWidget(self.search_input)
        self.search_layout.addWidget(self.search_button)
        self.search_layout.addWidget(self.menu_button)

        self.results_layout = QVBoxLayout()
        self.results_layout.addWidget(self.results_table)

        self.layout = QVBoxLayout()
        self.layout.addLayout(self.search_layout)
        self.layout.addLayout(self.results_layout)

        self.central_widget = QWidget()
        self.central_widget.setLayout(self.layout)
        self.setCentralWidget(self.central_widget)

        # Set custom item delegate for the search results table
        delegate = CustomItemDelegate(self.results_table)
        self.results_table.setItemDelegate(delegate)

    def get_excel_file_path(self):
        temp_path, _ = QFileDialog.getOpenFileName(self, 'Select Excel File', '', 'Excel Files (*.xlsx *.xls)')

        if temp_path:  # A file was selected
            self.excel_file_path = temp_path
            try:
                self.sheet_names = pd.read_excel(self.excel_file_path, sheet_name=None)  # Load the sheet names and data frames
                if len(self.sheet_names) == 1:  # Only one sheet available
                    sheet_name = list(self.sheet_names.keys())[0]  # Get the sheet name
                    self.excel_settings.sheet_name = sheet_name  # Update the selected sheet name
                    self.open_settings()  # Open the settings dialog directly
            except Exception as e:
                self.show_error_dialog('Error', str(e))

    def open_settings(self):
        if not self.excel_file_path:
            self.show_error_dialog('Error', 'No Excel file selected.')
            return

        dialog = ExcelSettingsDialog(self.excel_settings, self.sheet_names, parent=self)
        dialog.settings_updated.connect(self.update_excel_settings)
        dialog.exec_()

    def update_excel_settings(self, new_settings):
        self.excel_settings = new_settings

    def search_excel(self):
        if not self.excel_file_path:
            self.show_error_dialog('Error', 'No Excel file selected.')
            return

        search_phrase = self.search_input.text()
        if not search_phrase:
            self.show_error_dialog('Error', 'Please enter a search phrase.')
            return

        if self.search_thread is not None and self.search_thread.isRunning():
            self.show_error_dialog('Error', 'Search is already in progress.')
            return

        self.results_table.clearContents()
        self.results_table.setColumnCount(0)
        self.results_table.setRowCount(0)
        self.search_results = []

        self.search_thread = SearchThread(self.excel_file_path, self.excel_settings, search_phrase)
        self.search_thread.search_complete.connect(self.display_search_results)
        self.search_thread.start()

    def display_search_results(self, data):
        self.search_results = data

        if not data:
            self.show_message_box('No Results', 'No matching results found.')
            return

        # Get the columns to display based on the selected parameters
        display_columns = ['Sheet'] + [self.excel_settings.search_column] + self.excel_settings.additional_columns

        # Set the number of columns in the table
        self.results_table.setColumnCount(len(display_columns))

        # Set the column labels
        self.results_table.setHorizontalHeaderLabels(display_columns)

        # Set the number of rows in the table
        self.results_table.setRowCount(len(data))

        # Populate the table with the search results
        for row_index, row_data in enumerate(data):
            for col_index, column in enumerate(display_columns):
                item = row_data.get(column)
                table_item = QTableWidgetItem(str(item))
                self.results_table.setItem(row_index, col_index, table_item)

        self.results_table.resizeColumnsToContents()
        self.results_table.resizeRowsToContents()

    def export_to_excel(self):
        if not self.search_results:
            self.show_message_box('No Results', 'No search results to export.')
            return

        file_path, _ = QFileDialog.getSaveFileName(self, 'Save Excel File', '', 'Excel Files (*.xlsx)')

        if file_path:
            try:
                df = pd.DataFrame(self.search_results)
                df.columns = ['Sheet'] + [self.excel_settings.search_column] + self.excel_settings.additional_columns
                df = df.drop_duplicates()

                with pd.ExcelWriter(file_path) as writer:
                    df.to_excel(writer, index=False)

                self.show_message_box('Export Successful', 'Search results have been exported to an Excel file.')
            except Exception as e:
                self.show_error_dialog('Error', str(e))

    def show_error_dialog(self, title, message):
        QMessageBox.critical(self, title, message, QMessageBox.Ok)

    def show_message_box(self, title, message):
        QMessageBox.information(self, title, message, QMessageBox.Ok)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = ExcelSearchApp()
    window.show()
    sys.exit(app.exec_())

